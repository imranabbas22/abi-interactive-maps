#!/usr/bin/env python3
"""Parse the DURABILITY tab from skrux data sheet."""
import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\imran\OneDrive\Desktop\skrux data sheet.xlsx", data_only=True)
ws = wb["DURABILITY"]

print("=== ARMORS TABLE (Left side) ===")
print(f"{'#':>3} | {'Name':22s} | {'In-Game':22s} | {'Material':18s} | {'WORN min (GREEN)':>16s} | {'LIKE-NEW min (RED)':>18s}")
print("-"*105)
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=7, values_only=False), 1):
    vals = [cell.value for cell in row]
    colors = []
    for cell in row:
        c = ""
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            rgb = str(cell.fill.start_color.rgb)
            if "FFB6D7A8" in rgb: c = "GREEN"
            elif "FFEA9999" in rgb: c = "RED"
            elif "FF00FF00" in rgb: c = "BRIGHT_GREEN"
            elif "FFFFFF00" in rgb: c = "YELLOW"
            elif "FFFF9900" in rgb: c = "ORANGE"
            elif "FFFF0000" in rgb: c = "BRIGHT_RED"
            elif "FF073763" in rgb: c = "DARK_BLUE"
            elif "FF1F1F1F" in rgb or "FF0D0D0D" in rgb: c = "DARK_BG"
        colors.append(c)
    if any(v is not None for v in vals) and any(v is not None for v in vals[1:5]):
        g = f" {colors[4]}" if colors[4] else ""
        r = f" {colors[5]}" if colors[5] else ""
        print(f"{i:3d} | {str(vals[0] or ''):22s} | {str(vals[1] or ''):22s} | {str(vals[2] or ''):22s} | {str(vals[3] or ''):18s} | {str(vals[4] or ''):>6s}{g:16s} | {str(vals[5] or ''):>6s}{r:18s}")

print()
print("=== HELMETS TABLE (Right side) ===")
print(f"{'#':>3} | {'Name':22s} | {'In-Game':22s} | {'Material':18s} | {'WORN min (GREEN)':>16s} | {'LIKE-NEW min (RED)':>18s}")
print("-"*105)
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=8, max_col=13, values_only=False), 1):
    vals = [cell.value for cell in row]
    colors = []
    for cell in row:
        c = ""
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            rgb = str(cell.fill.start_color.rgb)
            if "FFB6D7A8" in rgb: c = "GREEN"
            elif "FFEA9999" in rgb: c = "RED"
            elif "FF00FF00" in rgb: c = "BRIGHT_GREEN"
            elif "FFFFFF00" in rgb: c = "YELLOW"
            elif "FFFF9900" in rgb: c = "ORANGE"
            elif "FFFF0000" in rgb: c = "BRIGHT_RED"
            elif "FF073763" in rgb: c = "DARK_BLUE"
            elif "FF1F1F1F" in rgb or "FF0D0D0D" in rgb: c = "DARK_BG"
        colors.append(c)
    if any(v is not None for v in vals) and any(v is not None for v in vals[1:4]):
        g = f" {colors[3]}" if colors[3] else ""
        r = f" {colors[4]}" if colors[4] else ""
        print(f"{i:3d} | {str(vals[0] or ''):22s} | {str(vals[1] or ''):22s} | {str(vals[2] or ''):18s} | {str(vals[3] or ''):>6s}{g:16s} | {str(vals[4] or ''):>6s}{r:18s}")

print()
print("=== MATERIAL FORMULAS (Bottom) ===")
for row in ws.iter_rows(min_row=30, max_row=ws.max_row, values_only=True):
    if any(v is not None for v in row[:7]):
        print(f"  {row}")

# Also print raw data for all rows to see location
print()
print("=== ALL RAW ROWS ===")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if any(v is not None for v in row):
        print(f"Row {i:2d}: {[(v, type(v).__name__) for v in row[:13] if v is not None]}")
